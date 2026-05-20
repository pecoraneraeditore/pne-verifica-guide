"""Genera il JSON dei ristoranti per la mappa auto-aggiornante.

v4: auto-detect di TUTTE le colonne tramite header (riga 7).
Diventa robusto a inserimenti/spostamenti di colonne nel xlsx.

Colonne riconosciute (header in riga 7, match case-insensitive):
  - "Nome"       (default col 4)
  - "Tipologia"  (default col 5)
  - "Indirizzo"  (default col 6)
  - "Zona"       (default col 7)
  - "Città"/"Citta" (default col 8)
  - "Recensore"  (default col 14)
  - "GPS"/"Coord*" (opzionale, override geocoding manuale)

Se un header non viene trovato, usa la posizione default originale.
"""
from __future__ import annotations
import argparse, json, re, sys, time, urllib.parse, urllib.request
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
import openpyxl

CATEGORIES = {
    'fatto':           {'label': 'Fatto',            'color': '#E53935'},
    'riservato':       {'label': 'Riservato per',    'color': '#757575'},
    'da_fare':         {'label': 'Da fare',          'color': '#FFFFFF'},
    'per_editori':     {'label': 'Per editori',      'color': '#F48FB1'},
    'da_finire':       {'label': 'Da finire',        'color': '#FBC02D'},
    'non_recensibile': {'label': 'Non recensibile',  'color': '#1565C0'},
    'altro_azzurro':   {'label': 'Altro',            'color': '#4FC3F7'},
    'plurimo':         {'label': 'Indirizzo plurimo','color': '#FB8C00'},
}

USER_AGENT = 'PNE-Mappa-Guida/4.0 (contatto: s.cargiani@lapecoranera.net)'

DEFAULT_COLS = {
    'nome': 4, 'tipologia': 5, 'indirizzo': 6, 'zona': 7,
    'citta': 8, 'recensore': 14, 'gps': None,
}


def categorize(fill):
    fg = fill.fgColor
    if fg.type == 'rgb':
        rgb = (fg.rgb or '').upper()
        if rgb == 'FFFF0000':           return 'fatto'
        if rgb == 'FFFFFF00':           return 'da_finire'
        if rgb == 'FFFFC000':           return 'plurimo'
        if rgb in ('00000000', '', None): return 'da_fare'
    elif fg.type == 'theme':
        t, tint = fg.theme, fg.tint or 0
        if t == 0 and tint < -0.2:    return 'riservato'
        if t == 5 and tint > 0.5:     return 'per_editori'
        if t == 3 and tint > 0.4:     return 'non_recensibile'
        if t == 4 and tint > 0.2:     return 'altro_azzurro'
    elif fg.type == 'indexed' and fg.indexed == 10:
        return 'fatto'
    return 'da_fare'


def trova_colonne(ws):
    """Mappa nome_logico -> indice colonna (1-based) leggendo gli header riga 7."""
    cols = dict(DEFAULT_COLS)
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=7, column=col).value
        if not val or not isinstance(val, str):
            continue
        v = val.lower().strip()
        if v == 'nome':                       cols['nome'] = col
        elif v == 'tipologia':                cols['tipologia'] = col
        elif v == 'indirizzo':                cols['indirizzo'] = col
        elif v == 'zona':                     cols['zona'] = col
        elif v in ('citta', 'città'):         cols['citta'] = col
        elif v == 'recensore':                cols['recensore'] = col
        elif 'gps' in v or 'coord' in v:      cols['gps'] = col
    return cols


def estrai_ristoranti(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb['Tavole e pause golose']
    cols = trova_colonne(ws)
    out = []
    for r in range(8, ws.max_row + 1):
        nome_cell = ws.cell(row=r, column=cols['nome'])
        nome = nome_cell.value
        if not nome or not str(nome).strip():
            continue
        rec = {
            'nome':      str(nome).strip(),
            'tipologia': str(ws.cell(row=r, column=cols['tipologia']).value or '').strip(),
            'indirizzo': str(ws.cell(row=r, column=cols['indirizzo']).value or '').strip(),
            'zona':      str(ws.cell(row=r, column=cols['zona']).value or '').strip(),
            'citta':     str(ws.cell(row=r, column=cols['citta']).value or '').strip(),
            'recensore': str(ws.cell(row=r, column=cols['recensore']).value or '').strip(),
            'categoria': categorize(nome_cell.fill),
        }
        rec['coord_gps'] = (str(ws.cell(row=r, column=cols['gps']).value or '').strip()
                            if cols['gps'] else '')
        out.append(rec)
    return out, cols


def parse_gps(s):
    if not s:
        return None
    s = s.strip().strip('()[]').replace(';', ',')
    parts = [p.strip() for p in s.split(',')]
    if len(parts) != 2:
        return None
    try:
        lat, lon = float(parts[0]), float(parts[1])
    except (ValueError, TypeError):
        return None
    if not (-90 <= lat <= 90 and -180 <= lon <= 180):
        return None
    if not (35 <= lat <= 48 and 6 <= lon <= 20):
        print(f"  ! coordinata sospetta (fuori Italia): {lat}, {lon}", file=sys.stderr)
    return lat, lon


_PREFISSI_VIA = (r'Via|Viale|V\.le|Piazza|P\.zza|P\.za|Corso|C\.so|Largo|Lgo'
                 r'|Vicolo|Strada|S\.da|Salita|Lungomare|Lungotevere|Borgo'
                 r'|Contrada|Localit[àa]|Loc\.?')

_CUT_MARKERS = [
    ' - Loc.', ' - Loc ', ' - loc.', ' - loc ',
    ' c/o ', ' C/O ',
    ' ang. ', ' Ang. ', ' angolo ',
    ' - Box ', ' Box ',
    ' Mercato ', ' mercato ',
    ' Galleria ', ' galleria ',
]


def normalizza(s):
    if not s:
        return ''
    s = s.replace('’', "'").replace('‘', "'")
    lower = s.lower()
    for cm in _CUT_MARKERS:
        idx = lower.find(cm.lower())
        if idx > 0:
            s = s[:idx]
            lower = s.lower()
    s = re.sub(r'(\d+)\s*/\s*[A-Za-z][\w\-/]*', r'\1', s)
    s = re.sub(r'(\d+)\s+[A-Z](?:/[A-Z])+\b', r'\1', s)
    s = re.sub(r',?\s*snc\s*$', '', s, flags=re.IGNORECASE)
    return s.strip(' ,')


def estrai_via(s):
    if not s:
        return ''
    pattern = rf'\b({_PREFISSI_VIA})\s+[^,]+(?:,\s*\d+\S*)?'
    m = re.search(pattern, s, flags=re.IGNORECASE)
    return m.group(0) if m else s


def via_senza_civico(s):
    return re.sub(r',\s*\d.*$', '', s).strip(' ,')


def cache_key(r):
    return f"{r['nome'].lower()}|{r['indirizzo'].lower()}|{r['citta'].lower()}"


def nominatim_query(q, timeout=10):
    url = ('https://nominatim.openstreetmap.org/search?'
           + urllib.parse.urlencode({
               'q': q, 'format': 'json', 'limit': '1', 'countrycodes': 'it',
           }))
    req = urllib.request.Request(url, headers={'User-Agent': USER_AGENT})
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            data = json.loads(resp.read())
    except Exception as e:
        print(f"  ! Nominatim error per '{q[:60]}...': {e}", file=sys.stderr)
        return None
    if not data:
        return None
    try:
        return float(data[0]['lat']), float(data[0]['lon'])
    except (KeyError, ValueError, IndexError):
        return None


def costruisci_varianti(r):
    ind = r['indirizzo']
    city = r['citta']
    varianti = []
    def add(q, approx):
        if q and (q, approx) not in varianti:
            varianti.append((q, approx))
    if ind and city: add(f"{ind}, {city}, Italia", False)
    elif ind:        add(f"{ind}, Italia", False)
    norm = normalizza(ind)
    if norm and city: add(f"{norm}, {city}, Italia", False)
    via = estrai_via(norm or ind)
    if via and city and via != norm: add(f"{via}, {city}, Italia", False)
    via_sc = via_senza_civico(via or norm or ind)
    if via_sc and city and via_sc != via: add(f"{via_sc}, {city}, Italia", False)
    if city: add(f"{city}, Italia", True)
    return varianti


def geocodifica_tutti(ristoranti, cache_path, no_geocode=False, sleep_s=1.1):
    cache = {}
    if cache_path.exists():
        try:
            cache = json.loads(cache_path.read_text(encoding='utf-8'))
        except json.JSONDecodeError:
            cache = {}
    n_cache, n_new, n_manual = 0, 0, 0
    falliti = []
    today = datetime.now(timezone.utc).strftime('%Y-%m-%d')
    for r in ristoranti:
        if r.get('coord_gps'):
            coords = parse_gps(r['coord_gps'])
            if coords:
                r['lat'], r['lon'] = coords
                r['approssimato'] = False
                r['manuale'] = True
                r['geocoded_at'] = today
                n_manual += 1
                continue
            else:
                print(f"  ! GPS non valido per '{r['nome']}': {r['coord_gps']!r}", file=sys.stderr)
        key = cache_key(r)
        if key in cache and cache[key].get('lat') is not None:
            r['lat'] = cache[key]['lat']
            r['lon'] = cache[key]['lon']
            r['approssimato'] = bool(cache[key].get('approssimato', False))
            r['manuale'] = False
            r['geocoded_at'] = cache[key].get('geocoded_at')
            n_cache += 1
            continue
        if no_geocode:
            r['lat'] = None
            r['lon'] = None
            falliti.append({'nome': r['nome'], 'motivo': 'no-geocode flag'})
            continue
        found = None
        approx_flag = False
        for q, approx in costruisci_varianti(r):
            coords = nominatim_query(q)
            time.sleep(sleep_s)
            if coords:
                found = coords
                approx_flag = approx
                break
        if found is None:
            r['lat'] = None
            r['lon'] = None
            falliti.append({
                'nome': r['nome'],
                'indirizzo': r['indirizzo'],
                'citta': r['citta'],
                'motivo': 'no result dopo cascata',
            })
            continue
        lat, lon = found
        r['lat'] = lat
        r['lon'] = lon
        r['approssimato'] = approx_flag
        r['manuale'] = False
        r['geocoded_at'] = today
        cache[key] = {
            'lat': lat, 'lon': lon,
            'approssimato': approx_flag, 'geocoded_at': today,
        }
        n_new += 1
        if approx_flag:
            print(f"  ~ {r['nome']}: approssimato a livello citta' ({r['citta']})")
    cache_path.parent.mkdir(parents=True, exist_ok=True)
    cache_path.write_text(
        json.dumps(cache, ensure_ascii=False, indent=2, sort_keys=True),
        encoding='utf-8',
    )
    return n_cache, n_new, n_manual, falliti


def scrivi_json(ristoranti, out_path, source_file, citta_default, regione, falliti, cols):
    out_path.parent.mkdir(parents=True, exist_ok=True)
    valid = [r for r in ristoranti if r.get('lat') is not None]
    for r in valid:
        r.pop('coord_gps', None)
    payload = {
        'generated_at': datetime.now(timezone.utc).isoformat(timespec='seconds'),
        'source_file': source_file,
        'citta_default': citta_default,
        'regione': regione,
        'categories': CATEGORIES,
        'restaurants': valid,
        'geocoding_failed': falliti,
        'stats': {
            'totale_letti':    len(ristoranti),
            'con_coordinate':  len(valid),
            'manuali':         sum(1 for r in valid if r.get('manuale')),
            'approssimati':    sum(1 for r in valid if r.get('approssimato')),
            'precisi':         sum(1 for r in valid if not r.get('approssimato') and not r.get('manuale')),
            'falliti':         len(falliti),
            'con_recensore':   sum(1 for r in ristoranti if r['recensore']),
            'colonne_rilevate': cols,
        },
    }
    out_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding='utf-8')


def main():
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument('xlsx', type=Path)
    p.add_argument('--out', type=Path, required=True)
    p.add_argument('--cache', type=Path, required=True)
    p.add_argument('--no-geocode', action='store_true')
    p.add_argument('--sleep', type=float, default=1.1)
    p.add_argument('--citta-default', default='Roma')
    p.add_argument('--regione', default='Lazio')
    args = p.parse_args()

    if not args.xlsx.exists():
        print(f"File non trovato: {args.xlsx}", file=sys.stderr)
        return 1

    print(f"Lettura {args.xlsx.name}...")
    ristoranti, cols = estrai_ristoranti(args.xlsx)
    cats = Counter(r['categoria'] for r in ristoranti)
    print(f"  letti {len(ristoranti)} ristoranti  categorie: {dict(cats)}")
    print(f"  colonne rilevate: {cols}")
    if cols['gps']:
        n_gps = sum(1 for r in ristoranti if r.get('coord_gps'))
        print(f"  -> colonna GPS attiva (col {cols['gps']}): {n_gps} ristoranti con coord manuali")

    print(f"Geocoding (cache: {args.cache.name})...")
    n_cache, n_new, n_manual, falliti = geocodifica_tutti(
        ristoranti, args.cache,
        no_geocode=args.no_geocode, sleep_s=args.sleep,
    )
    valid = [r for r in ristoranti if r.get('lat') is not None]
    n_manuali = sum(1 for r in valid if r.get('manuale'))
    n_approx = sum(1 for r in valid if r.get('approssimato'))
    n_precisi = sum(1 for r in valid if not r.get('approssimato') and not r.get('manuale'))
    print(f"  manuali (da GPS xlsx):       {n_manuali}")
    print(f"  precisi (geocode preciso):   {n_precisi}")
    print(f"  approssimati (liv. citta'):  {n_approx}")
    print(f"  falliti del tutto:           {len(falliti)}")

    scrivi_json(ristoranti, args.out, args.xlsx.name,
                args.citta_default, args.regione, falliti, cols)
    print(f"JSON scritto: {args.out}")
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
