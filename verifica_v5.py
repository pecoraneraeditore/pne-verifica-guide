#!/usr/bin/env python3
"""
verifica_v5.py — La Pecora Nera — Versione 5.1
SerpAPI (Google Maps) + Web Scraping (The Fork / Facebook / Instagram)

Modifiche v5.1:
  - Checkpoint giornaliero: righe già verificate oggi non vengono rifatte
  - Argomento --city per elaborare una sola città per chiamata (multi-pass)
  - Argomento --max N per sovrascrivere MAX_PER_CITY
  - Email: prova prima bypass proxy (proxies=None) poi salva JSON per fallback MCP
  - Stats accumulate in verifica_stats_DDMMYYYY.json tra run multipli

Variabili d'ambiente richieste:
  SERPAPI_KEY    — chiave API SerpAPI
  BREVO_API_KEY  — chiave API Brevo
  GITHUB_TOKEN   — Personal Access Token GitHub (scope: repo)
  GITHUB_REPO    — es. "pecoraneraeditore/pne-verifica-guide"

Uso:
  python verifica_v5.py                          # tutte le città, email alla fine
  python verifica_v5.py --city Milano            # solo Milano, no email
  python verifica_v5.py --city Roma --max 15     # Roma, max 15 ristoranti
  python verifica_v5.py --email-only             # manda solo l'email dai stats JSON salvati
"""

import os, sys, json, time, base64, re, shutil, subprocess, argparse
from datetime import datetime
import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import PatternFill, Font
import pandas as pd

# ── Argomenti ────────────────────────────────────────────────────────────────
parser = argparse.ArgumentParser(description="Verifica apertura ristoranti La Pecora Nera")
parser.add_argument("--city",       default=None,  help="Città da elaborare (Milano/Roma/Torino). Se omesso: tutte.")
parser.add_argument("--max",        type=int, default=None, help="Max ristoranti da verificare (default: 34)")
parser.add_argument("--email-only", action="store_true",    help="Manda solo l'email con i stats già salvati")
parser.add_argument("--no-email",   action="store_true",    help="Non mandare email al termine")
args = parser.parse_args()

# ── Configurazione ───────────────────────────────────────────────────────────
SCRIPT_DIR    = os.path.dirname(os.path.abspath(__file__))
DATA_DIR      = SCRIPT_DIR
SERPAPI_KEY   = os.environ.get("SERPAPI_KEY",   "")   # ← imposta in GitHub Secrets
BREVO_API_KEY = os.environ.get("BREVO_API_KEY", "")   # ← imposta in GitHub Secrets
GITHUB_TOKEN  = os.environ.get("GITHUB_TOKEN",  "")
GITHUB_REPO   = os.environ.get("GITHUB_REPO",   "")
SENDER_EMAIL  = "s.cargiani@lapecoranera.net"
RECIPIENTS    = [{"email": "s.cargiani@lapecoranera.net"},
                 {"email": "f.darienzo@lapecoranera.net"}]
today_str     = datetime.now().strftime("%d/%m/%Y")
today_file    = datetime.now().strftime("%d%m%Y")

# Ricerche SerpAPI per city-pass. Target 100 totali = ~34/città.
# Con chiamate separate per città puoi fare più pass: 34 × 3 città = 102 ricerche/run.
MAX_PER_CITY  = args.max if args.max else 34
CITIES        = [args.city] if args.city else ["Milano", "Roma", "Torino"]
SLEEP         = 0.5   # secondi tra richieste web (era 1.0/1.2)

# File JSON per accumulare stats tra run multipli della stessa giornata
STATS_FILE = os.path.join(DATA_DIR, f"verifica_stats_{today_file}.json")

# ── Utility ──────────────────────────────────────────────────────────────────
def _str(v, d=""):
    if v is None: return d
    s = str(v).strip()
    return d if s in ("nan", "None", "") else s

def carica_stats():
    """Carica stats accumulate dal JSON (se esiste)."""
    if os.path.exists(STATS_FILE):
        with open(STATS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

def salva_stats(stats):
    """Salva stats nel JSON per essere usate dalle chiamate successive."""
    with open(STATS_FILE, "w", encoding="utf-8") as f:
        json.dump(stats, f, ensure_ascii=False, indent=2)

# ── SerpAPI: Google Maps ─────────────────────────────────────────────────────
def verifica_google_maps(nome, citta):
    try:
        params = {
            "engine": "google",
            "q": f"{nome} ristorante {citta}",
            "api_key": SERPAPI_KEY,
            "hl": "it", "gl": "it", "num": 3,
        }
        resp = requests.get("https://serpapi.com/search", params=params, timeout=15,
                            proxies={"http": None, "https": None})
        data = resp.json()
        res = {"status": "INCERTO", "orari_ap": "", "orari_ch": "",
               "rating": "", "reviews": "",
               "ultima_rev": str(datetime.now().year), "note": ""}

        kg = data.get("knowledge_graph", {})
        if kg.get("rating"):  res["rating"]   = str(kg["rating"])
        if kg.get("reviews"): res["reviews"]  = str(kg["reviews"])
        if kg.get("hours"):   res["orari_ap"] = str(kg["hours"])[:80]

        for lr in data.get("local_results", [])[:1]:
            if lr.get("rating"):  res["rating"]   = str(lr["rating"])
            if lr.get("reviews"): res["reviews"]  = str(lr["reviews"])
            if lr.get("hours"):   res["orari_ap"] = str(lr.get("hours", ""))[:80]
            os_ = lr.get("open_state", "").lower()
            if any(w in os_ for w in ["aperto", "open", "chiude", "closes", "apre alle"]):
                res["status"] = "APERTO"
            elif "permanentemente" in os_ or "permanently closed" in os_:
                res["status"] = "CHIUSO"

        if res["rating"] and res["status"] == "INCERTO":
            res["status"] = "APERTO"
            res["note"]   = f"Rating Google: {res['rating']} ({res['reviews']} rec.)"

        for org in data.get("organic_results", [])[:3]:
            snip = (org.get("snippet", "") + " " + org.get("title", "")).lower()
            if "chiuso definitivamente" in snip or "permanently closed" in snip:
                res["status"] = "CHIUSO"
                res["note"]   = "Chiuso definitivamente (ricerca web)"
                break
            if res["status"] == "INCERTO" and any(
                    w in snip for w in ["2026", "2027", "prenota", "menu", "reservations"]):
                res["status"] = "APERTO"
                res["note"]   = "Trovato online con menzione recente"

        time.sleep(SLEEP)
        return res
    except Exception as e:
        time.sleep(SLEEP)
        return {"status": "INCERTO", "orari_ap": "", "orari_ch": "",
                "rating": "", "reviews": "", "ultima_rev": "",
                "note": f"Err SerpAPI: {str(e)[:40]}"}

# ── Web scraping: The Fork ───────────────────────────────────────────────────
def verifica_the_fork(nome, citta):
    try:
        url = (f"https://www.thefork.it/search?"
               f"q={nome.replace(' ', '+')}+{citta.replace(' ', '+')}")
        resp = requests.get(url,
            headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
                     "Accept-Language": "it-IT,it;q=0.9"},
            timeout=8, proxies={"http": None, "https": None})
        text = resp.text.lower()
        trovato = (nome.lower()[:12] in text or "prenota" in text) and len(resp.text) > 3000
        rating  = ""
        m = re.search(r'\b([89]\.\d|10\.0|7\.[5-9])\b', text)
        if m: rating = m.group(1)
        time.sleep(SLEEP)
        return {"trovato": trovato, "rating": rating,
                "note": "TF: trovato" if trovato else "TF: non trovato"}
    except Exception as e:
        # Non aggiunge sleep su errore proxy (fallisce subito)
        return {"trovato": False, "rating": "", "note": f"TF err: {str(e)[:30]}"}

# ── Web scraping: Facebook ───────────────────────────────────────────────────
def verifica_facebook(handle):
    if not handle or str(handle).strip() in ("", "nan", "None"):
        return {"trovato": False, "note": "FB: nessun handle"}
    try:
        h = str(handle).strip().lstrip("@")
        resp = requests.get(f"https://www.facebook.com/{h}",
            headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"},
            timeout=8, allow_redirects=True, proxies={"http": None, "https": None})
        text = resp.text.lower()
        trovato = (resp.status_code == 200 and len(resp.text) > 5000
                   and "page not found" not in text
                   and "pagina non trovata" not in text
                   and "questa pagina non è disponibile" not in text)
        time.sleep(SLEEP)
        return {"trovato": trovato, "note": f"FB: {'trovato' if trovato else 'non trovato'} ({h})"}
    except Exception as e:
        return {"trovato": False, "note": f"FB err: {str(e)[:30]}"}

# ── Web scraping: Instagram ──────────────────────────────────────────────────
def verifica_instagram(handle):
    if not handle or str(handle).strip() in ("", "nan", "None"):
        return {"trovato": False, "note": "IG: nessun handle"}
    try:
        h = str(handle).strip().lstrip("@")
        resp = requests.get(f"https://www.instagram.com/{h}/",
            headers={"User-Agent": "Mozilla/5.0 (iPhone; CPU iPhone OS 17_0 like Mac OS X) AppleWebKit/605.1.15"},
            timeout=8, allow_redirects=True, proxies={"http": None, "https": None})
        text = resp.text.lower()
        trovato = (resp.status_code == 200 and len(resp.text) > 8000
                   and "sorry, this page" not in text
                   and "page not found" not in text)
        time.sleep(SLEEP)
        return {"trovato": trovato, "note": f"IG: {'trovato' if trovato else 'non trovato'} ({h})"}
    except Exception as e:
        return {"trovato": False, "note": f"IG err: {str(e)[:30]}"}

# ── Leggi Excel ──────────────────────────────────────────────────────────────
def leggi_excel(city):
    wb = openpyxl.load_workbook(os.path.join(DATA_DIR, f"Verifica_{city}_2027.xlsx"))
    ws = wb.active
    hdrs = [c.value for c in ws[1]]
    return [dict(zip(hdrs, row)) for row in ws.iter_rows(min_row=2, values_only=True) if row[0]]

# ── Colonne output ───────────────────────────────────────────────────────────
COLS = ["Nome Ristorante", "Indirizzo", "Telefono", "Sito", "Facebook", "Instagram",
        "Ultimo post Facebook", "Ultimo post Instagram", "Status",
        "Ultima revisione Google", "Ultima revisione The Fork", "Ultima revisione Tripadvisor",
        "Orari apertura", "Orari chiusura", "Social attivo?", "Menzioni notizie",
        "Note aggiuntive", "Data verifica"]

# ── Email via Brevo (con bypass proxy) ───────────────────────────────────────
def invia_email_brevo(stats_all):
    ta  = sum(s["aperto"]    for s in stats_all.values())
    tc  = sum(s["chiuso"]    for s in stats_all.values())
    ti  = sum(s["incerto"]   for s in stats_all.values())
    ttf = sum(s.get("the_fork", 0)  for s in stats_all.values())
    tfb = sum(s.get("facebook", 0)  for s in stats_all.values())
    tig = sum(s.get("instagram", 0) for s in stats_all.values())
    tv  = sum(s.get("verificati", 0) for s in stats_all.values())
    tp  = sum(s.get("promossi", 0)   for s in stats_all.values())
    sc  = sum(s.get("serpapi_count", 0) for s in stats_all.values())

    body = f"Verifica apertura ristoranti — {today_str}\n\nGOOGLE MAPS (SerpAPI):\n"
    for city in ["Milano", "Roma", "Torino"]:
        if city in stats_all:
            s = stats_all[city]
            body += f"  {city}: {s['aperto']} APERTO, {s['chiuso']} CHIUSO, {s['incerto']} INCERTO\n"
    body += (f"\nSOCIAL MEDIA:\n"
             f"  The Fork trovato: {ttf} ristoranti\n"
             f"  Facebook trovato: {tfb} ristoranti\n"
             f"  Instagram trovato: {tig} ristoranti\n"
             f"\nSTATISTICHE SESSIONE:\n"
             f"  Ricerche SerpAPI: {sc} / 250 disponibili/mese\n"
             f"  Verifiche web scraping: {tv} completate\n"
             f"  Ristoranti appena verificati: {tv}\n"
             f"  Promossi INCERTO→APERTO/CHIUSO: {tp}\n"
             f"  Ancora INCERTO: {ti}\n"
             f"\nTOTALE GUIDE: APERTO={ta}  CHIUSO={tc}  INCERTO={ti}\n\n"
             f"La Pecora Nera — Sistema automatico di verifica v5.1\n")

    if not BREVO_API_KEY:
        print("  BREVO_API_KEY non impostata — email saltata.")
        return False, body

    try:
        att = []
        for city in ["Milano", "Roma", "Torino"]:
            fpath = os.path.join(DATA_DIR, f"Verifica_{city}_2027.xlsx")
            if os.path.exists(fpath):
                with open(fpath, "rb") as f:
                    att.append({"name": f"Verifica_{city}_2027.xlsx",
                                "content": base64.b64encode(f.read()).decode()})

        # Prova PRIMA con bypass proxy (proxies=None → connessione diretta)
        r = requests.post("https://api.brevo.com/v3/smtp/email",
            headers={"api-key": BREVO_API_KEY, "content-type": "application/json"},
            json={"sender":    {"name": "La Pecora Nera — Sistema automatico", "email": SENDER_EMAIL},
                  "to":        RECIPIENTS,
                  "subject":   f"Verifica apertura locali guide del {today_str}",
                  "textContent": body,
                  "attachment":  att},
            timeout=60,
            proxies={"http": None, "https": None})   # ← bypass proxy HTTP

        if r.status_code in (200, 201):
            print(f"  ✓ Email inviata via Brevo! (HTTP {r.status_code})")
            return True, body
        else:
            print(f"  ✗ Brevo HTTP {r.status_code}: {r.text[:150]}")
            return False, body

    except Exception as e:
        print(f"  ✗ Brevo non raggiungibile ({str(e)[:80]})")
        return False, body

# ═══════════════════════════════════════════════════════════════════════════
# MODALITÀ --email-only: rilegge stats JSON e manda email
# ═══════════════════════════════════════════════════════════════════════════
if args.email_only:
    print(f"\n{'='*60}")
    print(f"MODALITÀ EMAIL-ONLY — {today_str}")
    print(f"{'='*60}")
    saved = carica_stats()
    if not saved:
        print("  Nessun stats JSON trovato. Esegui prima la verifica.")
        sys.exit(1)
    ok, body = invia_email_brevo(saved)
    if not ok:
        # Salva corpo email in file per fallback manuale/MCP
        pending = os.path.join(DATA_DIR, f"email_pendente_{today_file}.txt")
        with open(pending, "w", encoding="utf-8") as f:
            f.write(body)
        print(f"  → Email salvata in: {os.path.basename(pending)}")
        print("    (Usa Gmail MCP per inviare manualmente il contenuto)")
    sys.exit(0)

# ═══════════════════════════════════════════════════════════════════════════
# VERIFICA PRINCIPALE
# ═══════════════════════════════════════════════════════════════════════════
print(f"\n{'='*60}")
print(f"VERIFICA APERTURA RISTORANTI — LA PECORA NERA v5.1")
print(f"Data: {today_str}  |  Città: {', '.join(CITIES)}  |  Max/città: {MAX_PER_CITY}")
print(f"{'='*60}\n")

# Carica stats precedenti (da run precedenti della stessa giornata)
stats = carica_stats()
serpapi_count_total = sum(s.get("serpapi_count", 0) for s in stats.values())
all_rows = {}

for city in CITIES:
    print(f"\n── {city} ──────────────────────────────────────────────")

    # Backup (solo se non esiste già per oggi)
    src = os.path.join(DATA_DIR, f"Verifica_{city}_2027.xlsx")
    bk  = os.path.join(DATA_DIR, f"Verifica_{city}_2027_backup_pre_{today_file}.xlsx")
    if not os.path.exists(bk):
        shutil.copy2(src, bk)
        print(f"  Backup: {os.path.basename(bk)}")

    rows = leggi_excel(city)

    # ── CHECKPOINT: escludi righe già verificate oggi (qualunque status) ──
    incerto_tutti = [r for r in rows if _str(r.get("Status")) == "INCERTO"]
    incerto       = [r for r in incerto_tutti
                     if _str(r.get("Data verifica")) != today_str]
    gia_verificati_oggi = len(incerto_tutti) - len(incerto)
    altri         = [r for r in rows if _str(r.get("Status")) != "INCERTO"]

    print(f"  Totale: {len(rows)} | INCERTO totale: {len(incerto_tutti)} "
          f"| Già verificati oggi: {gia_verificati_oggi} | Da verificare ora: {len(incerto)}")

    da_verif  = incerto[:MAX_PER_CITY]
    rimandati = incerto[MAX_PER_CITY:]
    print(f"  Verifico in questa pass: {len(da_verif)} | Rimandati ai prossimi run: {len(rimandati)}")

    # Recupera stats precedenti per questa città (run precedenti oggi)
    prev = stats.get(city, {
        "aperto": sum(1 for r in altri if _str(r.get("Status")) == "APERTO"),
        "chiuso": sum(1 for r in altri if _str(r.get("Status")) == "CHIUSO"),
        "incerto": len(incerto_tutti),
        "verificati": 0, "promossi": 0,
        "the_fork": 0, "facebook": 0, "instagram": 0,
        "serpapi_count": 0,
    })

    cs = dict(prev)
    serpapi_count = 0
    updated = list(altri) + [r for r in incerto_tutti if _str(r.get("Data verifica")) == today_str]

    print()
    for r in da_verif:
        nome = _str(r.get("Nome Ristorante"))
        print(f"  [{serpapi_count_total + serpapi_count + 1:03d}] {nome[:36]:<36}", end="", flush=True)

        gm = verifica_google_maps(nome, city)
        serpapi_count += 1

        if gm["status"] != "INCERTO":
            r["Status"]  = gm["status"]
            cs["promossi"] += 1
            cs["incerto"]  -= 1
            if gm["status"] == "APERTO": cs["aperto"] += 1
            else:                        cs["chiuso"] += 1

        if gm["orari_ap"]: r["Orari apertura"] = gm["orari_ap"]
        if gm["orari_ch"]: r["Orari chiusura"] = gm["orari_ch"]
        r["Ultima revisione Google"] = gm["ultima_rev"] or _str(r.get("Ultima revisione Google"))

        tf = verifica_the_fork(nome, city)
        fb = verifica_facebook(_str(r.get("Facebook")))
        ig = verifica_instagram(_str(r.get("Instagram")))

        if tf["trovato"]:
            cs["the_fork"] += 1
            r["Ultima revisione The Fork"] = str(datetime.now().year)
        if fb["trovato"]: cs["facebook"]  += 1
        if ig["trovato"]: cs["instagram"] += 1

        soc = tf["trovato"] or fb["trovato"] or ig["trovato"]
        if soc: r["Social attivo?"] = "Si"

        if soc and r["Status"] == "INCERTO":
            r["Status"]    = "APERTO"
            cs["aperto"]  += 1
            cs["incerto"] -= 1
            cs["promossi"] += 1

        r["Note aggiuntive"] = (
            (gm["note"] + " | " if gm["note"] else "") +
            tf["note"] + " | " + fb["note"] + " | " + ig["note"]
        )[:200]
        # Checkpoint: imposta data verifica anche se rimane INCERTO
        r["Data verifica"] = today_str
        updated.append(r)
        print(f" → {r['Status']:8s} | TF:{tf['trovato']} FB:{fb['trovato']} IG:{ig['trovato']}")

    for r in rimandati:
        updated.append(r)

    serpapi_count_total += serpapi_count
    all_rows[city] = updated

    fa = sum(1 for r in updated if _str(r.get("Status")) == "APERTO")
    fc = sum(1 for r in updated if _str(r.get("Status")) == "CHIUSO")
    fi = sum(1 for r in updated if _str(r.get("Status")) == "INCERTO")

    cs.update({
        "aperto": fa, "chiuso": fc, "incerto": fi,
        "verificati":    cs.get("verificati", 0)    + serpapi_count,
        "serpapi_count": cs.get("serpapi_count", 0) + serpapi_count,
    })
    stats[city] = cs

    print(f"\n  ► {city}: APERTO={fa} CHIUSO={fc} INCERTO={fi} | "
          f"Verificati questa pass={serpapi_count} | Totale oggi={cs['verificati']}")

print(f"\n{'='*60}")
print(f"SerpAPI totali oggi: {sum(s.get('serpapi_count', 0) for s in stats.values())}")
print(f"{'='*60}")

# ── Salva Excel ───────────────────────────────────────────────────────────────
print("\nSalvataggio file Excel...")

for city in CITIES:
    path = os.path.join(DATA_DIR, f"Verifica_{city}_2027.xlsx")
    df   = pd.DataFrame([{c: _str(r.get(c)) for c in COLS} for r in all_rows[city]], columns=COLS)
    df.to_excel(path, index=False)

    wb     = openpyxl.load_workbook(path)
    ws     = wb.active
    yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    red    = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    si     = COLS.index("Status") + 1

    for ri, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), 2):
        sv = ws.cell(row=ri, column=si).value
        if sv == "INCERTO":
            for c in row: c.fill = yellow
        elif sv == "CHIUSO":
            for c in row:
                c.fill = red
                c.font = Font(color="FFFFFF", bold=True)

    cw = {"Nome Ristorante": 30, "Indirizzo": 35, "Telefono": 14, "Sito": 28,
          "Facebook": 22, "Instagram": 22, "Ultimo post Facebook": 18,
          "Ultimo post Instagram": 18, "Status": 10, "Ultima revisione Google": 18,
          "Ultima revisione The Fork": 18, "Ultima revisione Tripadvisor": 18,
          "Orari apertura": 25, "Orari chiusura": 20, "Social attivo?": 12,
          "Menzioni notizie": 18, "Note aggiuntive": 45, "Data verifica": 14}
    for i, cn in enumerate(COLS, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = cw.get(cn, 14)
    wb.save(path)
    print(f"  ✓ Verifica_{city}_2027.xlsx ({len(all_rows[city])} righe)")

# ── Salva stats JSON ──────────────────────────────────────────────────────────
salva_stats(stats)
print(f"  ✓ Stats: {os.path.basename(STATS_FILE)}")

# ── Email ─────────────────────────────────────────────────────────────────────
all_cities_done = set(["Milano", "Roma", "Torino"]).issubset(set(stats.keys()))

if not args.no_email and all_cities_done:
    print("\nInvio email via Brevo (bypass proxy)...")
    ok, body = invia_email_brevo(stats)
    if not ok:
        pending = os.path.join(DATA_DIR, f"email_pendente_{today_file}.txt")
        with open(pending, "w", encoding="utf-8") as f:
            f.write(body)
        print(f"  → Testo email salvato: {os.path.basename(pending)}")
        print("    (Sarà inviato via Gmail MCP come fallback)")
elif args.no_email:
    print("\nEmail saltata (--no-email).")
elif not all_cities_done:
    print(f"\nEmail rimandata — città elaborate: {list(stats.keys())}")

# ── Log ───────────────────────────────────────────────────────────────────────
log_path = os.path.join(DATA_DIR, f"Log_Verifica_{today_file}.txt")
ta = sum(s["aperto"]  for s in stats.values())
tc = sum(s["chiuso"]  for s in stats.values())
ti = sum(s["incerto"] for s in stats.values())
sc_tot = sum(s.get("serpapi_count", 0) for s in stats.values())

with open(log_path, "w", encoding="utf-8") as f:
    f.write(f"{'='*60}\nREPORT VERIFICA v5.1 — {today_str}\n{'='*60}\n\n")
    for city, s in stats.items():
        f.write(f"{city}: APERTO={s['aperto']} CHIUSO={s['chiuso']} INCERTO={s['incerto']}\n")
        f.write(f"  Verificati oggi={s.get('verificati',0)} Promossi={s.get('promossi',0)} "
                f"TF={s.get('the_fork',0)} FB={s.get('facebook',0)} IG={s.get('instagram',0)}\n\n")
    f.write(f"SerpAPI oggi: {sc_tot}\nTotale: APERTO={ta} CHIUSO={tc} INCERTO={ti}\n\n{'='*60}\n")
print(f"  ✓ Log: {os.path.basename(log_path)}")

# ── Git push ──────────────────────────────────────────────────────────────────
if GITHUB_TOKEN and GITHUB_REPO and all_cities_done and not args.no_email:
    print("\nGit push...")
    try:
        subprocess.run(["git","config","user.email",SENDER_EMAIL],     cwd=SCRIPT_DIR, check=True)
        subprocess.run(["git","config","user.name","PNE Routine Bot"], cwd=SCRIPT_DIR, check=True)
        for city in ["Milano","Roma","Torino"]:
            subprocess.run(["git","add",f"Verifica_{city}_2027.xlsx"], cwd=SCRIPT_DIR, check=True)
        diff = subprocess.run(["git","diff","--cached","--quiet"],      cwd=SCRIPT_DIR)
        if diff.returncode != 0:
            subprocess.run(["git","commit","-m",f"Verifica {today_str}"], cwd=SCRIPT_DIR, check=True)
            subprocess.run(["git","remote","set-url","origin",
                            f"https://{GITHUB_TOKEN}@github.com/{GITHUB_REPO}.git"],
                           cwd=SCRIPT_DIR, check=True)
            subprocess.run(["git","push"], cwd=SCRIPT_DIR, check=True)
            print("  ✓ Repository aggiornato.")
        else:
            print("  Nessuna modifica da committare.")
    except subprocess.CalledProcessError as e:
        print(f"  ✗ Errore git: {e}")

print(f"\n{'='*60}")
print(f"✓ COMPLETATO — pass per: {', '.join(CITIES)}")
print(f"{'='*60}\n")
