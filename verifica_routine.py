#!/usr/bin/env python3
"""
verifica_routine.py  –  Verifica apertura ristoranti guide La Pecora Nera
Progettato per girare come GitHub Actions (cloud) oppure localmente.
Variabili d'ambiente richieste:
  BREVO_API_KEY           – chiave API Brevo (invio email)
  GITHUB_TOKEN            – Personal Access Token GitHub (scope: repo)
  GITHUB_REPO             – es. "pecoraneraeditore/pne-verifica-guide"
  DROPBOX_APP_KEY         – App key dell'app Dropbox (facoltativa)
  DROPBOX_APP_SECRET      – App secret dell'app Dropbox (facoltativa)
  DROPBOX_REFRESH_TOKEN   – refresh token Dropbox (facoltativo)
Se le 3 variabili DROPBOX_* sono tutte impostate, lo script scarica i
Verifica_*.xlsx più aggiornati da Dropbox prima di processare, e ricarica
i risultati su Dropbox al termine. Se mancano, usa le versioni del repo.
"""
import os, sys, re, base64, subprocess, asyncio, json
from datetime import datetime, timedelta
# ── auto-install dipendenze ────────────────────────────────────────────────
def _ensure(pkg):
    try: __import__(pkg)
    except ImportError:
        subprocess.run([sys.executable, "-m", "pip", "install", pkg, "-q"], check=True)
for _p in ["pandas", "openpyxl", "requests", "playwright"]:
    _ensure(_p)
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font
import requests
PLAYWRIGHT_AVAILABLE = False
try:
    from playwright.async_api import async_playwright, TimeoutError as AsyncPWTimeout
    PLAYWRIGHT_AVAILABLE = True
except ImportError:
    print("Playwright non disponibile — social media saltati.")
# ── configurazione ─────────────────────────────────────────────────────────
SCRIPT_DIR    = os.path.dirname(os.path.abspath(__file__))
DATA_DIR      = SCRIPT_DIR
BREVO_API_KEY = os.environ.get("BREVO_API_KEY", "")
GITHUB_TOKEN  = os.environ.get("GITHUB_TOKEN", "")
GITHUB_REPO   = os.environ.get("GITHUB_REPO", "")
DROPBOX_APP_KEY       = os.environ.get("DROPBOX_APP_KEY", "")
DROPBOX_APP_SECRET    = os.environ.get("DROPBOX_APP_SECRET", "")
DROPBOX_REFRESH_TOKEN = os.environ.get("DROPBOX_REFRESH_TOKEN", "")
DROPBOX_FOLDER        = "/PNE Simone/Guide"
SENDER_EMAIL  = "s.cargiani@lapecoranera.net"
RECIPIENTS    = [{"email": "s.cargiani@lapecoranera.net"},
                 {"email": "f.darienzo@lapecoranera.net"}]
today      = datetime.now()
today_str  = today.strftime("%d/%m/%Y")
cutoff_30d = today - timedelta(days=30)
MAX_SOCIAL   = 300   # handle massimi da controllare per run
CONCURRENCY  = 3     # pagine parallele
TIMEOUT_MS   = 8000  # timeout per pagina
SOCIAL_DAYS  = 40    # giorni entro cui un post social conta come "attività recente"
CONFIRMED_APERTO = {
    "Ristorante D O":  {"note": "Michelin 2026 confermato (Cornaredo)",
                        "orari_ap": "12:00", "orari_ch": "14:30, 20:00-22:30"},
    "Erba Brusca":     {"note": "Confermato aperto feb 2026 (Yelp/sito)",
                        "orari_ap": "12:00 (gio-dom), 20:00 (mer-dom)", "orari_ch": "14:00, 22:30"},
    "Remulass":        {"note": "Guida Michelin 2026 confermata", "orari_ap": "", "orari_ch": ""},
    "Frangente":       {"note": "Listato tra migliori Milano 2026 (Gambero Rosso)", "orari_ap": "", "orari_ch": ""},
    "Cucina Franca":   {"note": "Presente su TheFork 2026", "orari_ap": "", "orari_ch": ""},
    "Bistrot 64":      {"note": "Aggiornato marzo 2026 (Yelp): lun 19:30-23:30, mar-sab 12-15:30 e 19:30-23:30",
                        "orari_ap": "12:00 (mar-sab), 19:30", "orari_ch": "15:30, 23:30"},
    "Glass Hostaria":  {"note": "Michelin Guide Roma 2026 confermato", "orari_ap": "", "orari_ch": ""},
    "Connubio":        {"note": "Confermato aperto (recensioni nov 2025, sito attivo)",
                        "orari_ap": "19:30", "orari_ch": "23:00"},
}
# ══════════════════════════════════════════════════════════════════════════
# DROPBOX HELPERS
# ══════════════════════════════════════════════════════════════════════════
def dropbox_get_access_token():
    """Ottiene un access token breve tramite il refresh token. None se non configurato."""
    if not (DROPBOX_APP_KEY and DROPBOX_APP_SECRET and DROPBOX_REFRESH_TOKEN):
        return None
    try:
        resp = requests.post(
            "https://api.dropboxapi.com/oauth2/token",
            auth=(DROPBOX_APP_KEY, DROPBOX_APP_SECRET),
            data={"refresh_token": DROPBOX_REFRESH_TOKEN, "grant_type": "refresh_token"},
            timeout=30,
        )
        resp.raise_for_status()
        return resp.json().get("access_token")
    except Exception as e:
        print(f"  Dropbox token error: {e}")
        return None
def dropbox_download(access_token, remote_path, local_path):
    """Scarica un file da Dropbox. Ritorna True/False."""
    try:
        resp = requests.post(
            "https://content.dropboxapi.com/2/files/download",
            headers={
                "Authorization": f"Bearer {access_token}",
                "Dropbox-API-Arg": json.dumps({"path": remote_path}),
            },
            timeout=60,
        )
        if resp.status_code == 200:
            with open(local_path, "wb") as f:
                f.write(resp.content)
            return True
        print(f"  Dropbox download fail {resp.status_code}: {remote_path}")
        return False
    except Exception as e:
        print(f"  Dropbox download exception for {remote_path}: {e}")
        return False
def dropbox_upload(access_token, local_path, remote_path):
    """Carica un file su Dropbox sovrascrivendo l'esistente. Ritorna True/False."""
    try:
        with open(local_path, "rb") as f:
            data = f.read()
        resp = requests.post(
            "https://content.dropboxapi.com/2/files/upload",
            headers={
                "Authorization": f"Bearer {access_token}",
                "Dropbox-API-Arg": json.dumps({
                    "path": remote_path, "mode": "overwrite",
                    "autorename": False, "mute": True,
                }),
                "Content-Type": "application/octet-stream",
            },
            data=data,
            timeout=120,
        )
        if resp.status_code == 200:
            return True
        print(f"  Dropbox upload fail {resp.status_code}: {remote_path}")
        return False
    except Exception as e:
        print(f"  Dropbox upload exception for {remote_path}: {e}")
        return False
# ══════════════════════════════════════════════════════════════════════════
# ASYNC SOCIAL MEDIA SCRAPING
# ══════════════════════════════════════════════════════════════════════════
def _parse_ts(ts):
    try: return datetime.fromtimestamp(int(ts)).strftime("%d/%m/%Y")
    except: return None
def _parse_iso(s):
    try: return datetime.fromisoformat(s.replace("Z","+00:00")).strftime("%d/%m/%Y")
    except: return None
async def _scrape_fb(handle, page):
    h = handle.strip().lstrip("@")
    try:
        await page.goto(f"https://m.facebook.com/{h}",
                        timeout=TIMEOUT_MS, wait_until="domcontentloaded")
        await page.wait_for_timeout(2000)
        for txt in ["Accetta tutto","Accetta tutti i cookie","Accetta","OK"]:
            try:
                await page.click(f'button:has-text("{txt}")', timeout=1200)
                await page.wait_for_timeout(600)
                break
            except: pass
        ts_list = await page.evaluate("""
            () => Array.from(document.querySelectorAll('abbr[data-utime]'))
                       .map(e => parseInt(e.getAttribute('data-utime')))
                       .filter(t => !isNaN(t))
        """)
        if ts_list:
            r = _parse_ts(max(ts_list))
            if r: return r
        dt_list = await page.evaluate("""
            () => Array.from(document.querySelectorAll('time[datetime]'))
                       .map(e => e.getAttribute('datetime'))
        """)
        dates = [_parse_iso(d) for d in dt_list if d]
        dates = sorted([d for d in dates if d])
        if dates: return dates[-1]
        content = await page.content()
        matches = re.findall(r'"publish_time":(\d{10})', content)
        if matches:
            r = _parse_ts(max(int(m) for m in matches))
            if r: return r
        return "N/D"
    except AsyncPWTimeout: return "Timeout"
    except Exception as e: return f"Err:{str(e)[:20]}"
async def _scrape_ig(handle, page):
    h = handle.strip().lstrip("@")
    try:
        await page.goto(f"https://www.instagram.com/{h}/",
                        timeout=TIMEOUT_MS, wait_until="domcontentloaded")
        await page.wait_for_timeout(2000)
        content = await page.content()
        matches = re.findall(r'"taken_at_timestamp"\s*:\s*(\d+)', content)
        if matches:
            r = _parse_ts(max(int(m) for m in matches))
            if r: return r
        matches2 = re.findall(r'"timestamp"\s*:\s*(\d{10})', content)
        if matches2:
            r = _parse_ts(max(int(m) for m in matches2))
            if r: return r
        dt_list = await page.evaluate("""
            () => Array.from(document.querySelectorAll('time[datetime]'))
                       .map(e => e.getAttribute('datetime'))
        """)
        dates = [_parse_iso(d) for d in dt_list if d]
        dates = sorted([d for d in dates if d])
        if dates: return dates[-1]
        return "N/D"
    except AsyncPWTimeout: return "Timeout"
    except Exception as e: return f"Err:{str(e)[:20]}"
async def scrape_social_batch(handles_list):
    """
    handles_list: [(key, "fb"|"ig", handle), ...]
    Ritorna: {key: date_string}
    """
    if not handles_list:
        return {}
    results = {}
    semaphore = asyncio.Semaphore(CONCURRENCY)
    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=True,
            args=["--no-sandbox","--disable-dev-shm-usage","--disable-gpu"]
        )
        async def scrape_one(key, platform, handle):
            async with semaphore:
                ctx = await browser.new_context(
                    user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
                    locale="it-IT",
                    viewport={"width": 1280, "height": 800}
                )
                page = await ctx.new_page()
                try:
                    if platform == "fb":
                        results[key] = await _scrape_fb(handle, page)
                    else:
                        results[key] = await _scrape_ig(handle, page)
                except Exception as e:
                    results[key] = f"Err:{str(e)[:20]}"
                finally:
                    await page.close()
                    await ctx.close()
        await asyncio.gather(*[scrape_one(k, pl, h) for k, pl, h in handles_list])
        await browser.close()
    return results
# ══════════════════════════════════════════════════════════════════════════
# UTILITY
# ══════════════════════════════════════════════════════════════════════════
def _str(v, default=""):
    if v is None: return default
    s = str(v).strip()
    return default if s in ("nan","None","") else s
# ══════════════════════════════════════════════════════════════════════════
# 0. DROPBOX SYNC (IN) — scarica le ultime Verifica_*.xlsx dal cloud
# ══════════════════════════════════════════════════════════════════════════
print("\n=== DROPBOX SYNC (IN) ===")
dbx_token = dropbox_get_access_token()
if dbx_token:
    for city in ["Milano", "Roma", "Torino"]:
        fname  = f"Verifica_{city}_2027.xlsx"
        remote = f"{DROPBOX_FOLDER}/{fname}"
        local  = os.path.join(DATA_DIR, fname)
        if dropbox_download(dbx_token, remote, local):
            print(f"  {fname} scaricato da Dropbox")
        else:
            print(f"  {fname} non scaricato — uso la versione del repo")
else:
    print("  Dropbox non configurato — uso le versioni del repo")
# ══════════════════════════════════════════════════════════════════════════
# 1. ELABORAZIONE PER CITTÀ  (prima passata — senza social)
# ══════════════════════════════════════════════════════════════════════════
rows_by_city   = {}   # {city: [row_dict, ...]}
stats          = {}
handles_needed = []   # [(key, "fb"|"ig", handle), ...]  — da passare ad async
for city in ["Milano", "Roma", "Torino"]:
    print(f"\n=== {city} ===")
    df_guide = pd.read_excel(os.path.join(DATA_DIR, f"{city}_2027.xlsx"), header=6)
    col_pl   = df_guide.columns[0]
    df_guide = df_guide[df_guide[col_pl] == 0].copy()
    df_guide = df_guide[df_guide["Nome"].notna()]
    df_guide = df_guide[df_guide["Nome"].astype(str).str.strip().str.len() > 0]
    cols     = ["Nome"] + [c for c in ["Indirizzo","Telefono"] if c in df_guide.columns]
    df_guide = df_guide[cols].copy()
    try:
        df_dati      = pd.read_excel(os.path.join(DATA_DIR, f"{city}_dati_2026_per_2027.xlsx"))
        nome_col     = next((c for c in df_dati.columns if c.lower() in ["nome","ristorante"]), None)
        internet_col = next((c for c in df_dati.columns if "internet" in c.lower() or "sito" in c.lower()), None)
        fb_col       = next((c for c in df_dati.columns if "facebook" in c.lower()), None)
        ig_col       = next((c for c in df_dati.columns if c.lower() == "instagram"), None)
        pick         = [c for c in [nome_col, internet_col, fb_col, ig_col] if c]
        if nome_col and internet_col and fb_col and ig_col:
            df_dati_clean = df_dati[pick].copy()
            df_dati_clean.columns = ["Nome","Sito","Facebook","Instagram"]
            df_dati_clean = df_dati_clean.drop_duplicates("Nome")
        else:
            df_dati_clean = pd.DataFrame(columns=["Nome","Sito","Facebook","Instagram"])
    except Exception as e:
        print(f"  Dati errore: {e}")
        df_dati_clean = pd.DataFrame(columns=["Nome","Sito","Facebook","Instagram"])
    verif_file = os.path.join(DATA_DIR, f"Verifica_{city}_2027.xlsx")
    if os.path.exists(verif_file):
        df_verif = pd.read_excel(verif_file)
        def _pd(d):
            try: return datetime.strptime(str(d).strip(), "%d/%m/%Y")
            except: return None
        df_verif["_date"] = df_verif["Data verifica"].apply(_pd)
        keep_mask   = df_verif["Status"].isin(["APERTO","CHIUSO"]) & (df_verif["_date"] >= cutoff_30d)
        df_keep     = df_verif[keep_mask].copy()
        df_reverify = df_verif[~keep_mask].copy()
    else:
        df_keep = df_reverify = pd.DataFrame()
    print(f"  Guide:{len(df_guide)} Mantieni:{len(df_keep)} Ri-verifica:{len(df_reverify)}")
    df_merged    = df_guide.merge(df_dati_clean, on="Nome", how="left")
    rows         = []
    new_verified = 0
    for idx, row in df_merged.iterrows():
        nome      = _str(row["Nome"])
        indirizzo = _str(row.get("Indirizzo"))
        telefono  = _str(row.get("Telefono"))
        sito      = _str(row.get("Sito"))
        facebook  = _str(row.get("Facebook"))
        instagram = _str(row.get("Instagram"))
        kept = df_keep[df_keep["Nome Ristorante"] == nome] if len(df_keep) else pd.DataFrame()
        if len(kept) > 0:
            k = kept.iloc[0]
            status=k["Status"]; data_v=k["Data verifica"]
            ul_google=_str(k.get("Ultima revisione Google")); ul_fork=_str(k.get("Ultima revisione The Fork"))
            ul_trip=_str(k.get("Ultima revisione Tripadvisor")); orari_ap=_str(k.get("Orari apertura"))
            orari_ch=_str(k.get("Orari chiusura")); social_att=_str(k.get("Social attivo?"))
            menzioni=_str(k.get("Menzioni notizie")); note=_str(k.get("Note aggiuntive"))
            fb_post=_str(k.get("Ultimo post Facebook")); ig_post=_str(k.get("Ultimo post Instagram"))
            if not sito:      sito      = _str(k.get("Sito"))
            if not facebook:  facebook  = _str(k.get("Facebook"))
            if not instagram: instagram = _str(k.get("Instagram"))
        elif nome in CONFIRMED_APERTO:
            conf = CONFIRMED_APERTO[nome]
            status="APERTO"; data_v=today_str
            ul_google=ul_fork=ul_trip=""
            orari_ap=conf.get("orari_ap",""); orari_ch=conf.get("orari_ch","")
            social_att="Si"; menzioni=""; note=conf.get("note","")
            fb_post=""; ig_post=""
            new_verified += 1
        else:
            prev = df_reverify[df_reverify["Nome Ristorante"] == nome] if len(df_reverify) else pd.DataFrame()
            if len(prev) > 0:
                p = prev.iloc[0]
                status="INCERTO"; data_v=today_str
                ul_google=_str(p.get("Ultima revisione Google")); ul_fork=_str(p.get("Ultima revisione The Fork"))
                ul_trip=_str(p.get("Ultima revisione Tripadvisor")); orari_ap=_str(p.get("Orari apertura"))
                orari_ch=_str(p.get("Orari chiusura")); social_att=_str(p.get("Social attivo?"))
                menzioni=_str(p.get("Menzioni notizie")); note="Non verificato in questa sessione"
                fb_post=_str(p.get("Ultimo post Facebook")); ig_post=_str(p.get("Ultimo post Instagram"))
                if not facebook:  facebook  = _str(p.get("Facebook"))
                if not instagram: instagram = _str(p.get("Instagram"))
            else:
                status="INCERTO"; data_v=today_str
                ul_google=ul_fork=ul_trip=orari_ap=orari_ch=social_att=menzioni=""
                note="Nuovo ristorante - da verificare"
                fb_post=""; ig_post=""
        row_dict = {
            "Nome Ristorante": nome, "Indirizzo": indirizzo, "Telefono": telefono,
            "Sito": sito, "Facebook": facebook, "Instagram": instagram,
            "Ultimo post Facebook": fb_post, "Ultimo post Instagram": ig_post,
            "Status": status, "Ultima revisione Google": ul_google,
            "Ultima revisione The Fork": ul_fork, "Ultima revisione Tripadvisor": ul_trip,
            "Orari apertura": orari_ap, "Orari chiusura": orari_ch,
            "Social attivo?": social_att, "Menzioni notizie": menzioni,
            "Note aggiuntive": note, "Data verifica": data_v,
        }
        rows.append(row_dict)
        # Raccogli handle da verificare (priorità: nessuna data esistente)
        row_key = f"{city}__{len(rows)-1}"
        if facebook and not fb_post:
            handles_needed.append((f"{row_key}__fb", "fb", facebook))
        if instagram and not ig_post:
            handles_needed.append((f"{row_key}__ig", "ig", instagram))
    rows_by_city[city] = rows
    stats[city] = {
        "totale": len(rows), "aperto": sum(1 for r in rows if r["Status"]=="APERTO"),
        "chiuso": sum(1 for r in rows if r["Status"]=="CHIUSO"),
        "incerto": sum(1 for r in rows if r["Status"]=="INCERTO"),
        "mantenuti": len(df_keep), "verificati_oggi": new_verified,
    }
    print(f"  Output:{len(rows)} APERTO={stats[city]['aperto']} CHIUSO={stats[city]['chiuso']} INCERTO={stats[city]['incerto']}")
# ══════════════════════════════════════════════════════════════════════════
# 2. SOCIAL MEDIA SCRAPING (async, max 300, distribuiti per città)
# ══════════════════════════════════════════════════════════════════════════
social_results = {}
if PLAYWRIGHT_AVAILABLE and handles_needed:
    # Distribuzione proporzionale per città
    per_city    = MAX_SOCIAL // 3
    selected    = []
    city_counts = {c: 0 for c in ["Milano","Roma","Torino"]}
    for key, platform, handle in handles_needed:
        city = key.split("__")[0]
        if city_counts[city] < per_city:
            selected.append((key, platform, handle))
            city_counts[city] += 1
    # Se una città ha meno di per_city, prendi il resto dalle altre
    remaining = MAX_SOCIAL - len(selected)
    if remaining > 0:
        for key, platform, handle in handles_needed:
            if (key, platform, handle) not in selected and remaining > 0:
                selected.append((key, platform, handle))
                remaining -= 1
    print(f"\n=== SOCIAL MEDIA SCRAPING ===")
    print(f"Handle da verificare: {len(selected)} / {len(handles_needed)} totali")
    print(f"Distribuzione: " + ", ".join(f"{c}={v}" for c,v in city_counts.items()))
    social_results = asyncio.run(scrape_social_batch(selected))
    nd_count  = sum(1 for v in social_results.values() if v == "N/D")
    ok_count  = sum(1 for v in social_results.values() if v and v[0].isdigit())
    err_count = len(social_results) - nd_count - ok_count
    print(f"Risultati: {ok_count} date trovate, {nd_count} N/D, {err_count} errori")
    # Mappa i risultati nelle righe
    for key, date in social_results.items():
        parts    = key.split("__")
        city     = parts[0]
        row_idx  = int(parts[1])
        platform = parts[2]
        if platform == "fb":
            rows_by_city[city][row_idx]["Ultimo post Facebook"]  = date
        else:
            rows_by_city[city][row_idx]["Ultimo post Instagram"] = date
    # Aggiorna contatori social nei stats
    for city in ["Milano","Roma","Torino"]:
        checked = sum(1 for k in social_results if k.startswith(city))
        stats[city]["social_checked"] = checked
else:
    for city in ["Milano","Roma","Torino"]:
        stats[city]["social_checked"] = 0
    if not PLAYWRIGHT_AVAILABLE:
        print("\nPlaywright non disponibile — social saltati.")
    else:
        print("\nNessun handle social da verificare.")
# ══════════════════════════════════════════════════════════════════════════
# 3. PROMOZIONE AUTOMATICA BASATA SU ATTIVITÀ SOCIAL RECENTE
#    INCERTO → APERTO se ha un post social negli ultimi 40 giorni
# ══════════════════════════════════════════════════════════════════════════
cutoff_social = today - timedelta(days=SOCIAL_DAYS)
promoted_total = 0
for city in ["Milano", "Roma", "Torino"]:
    promoted = 0
    for row in rows_by_city[city]:
        if row["Status"] != "INCERTO":
            continue
        for col in ["Ultimo post Facebook", "Ultimo post Instagram"]:
            date_str = row.get(col, "")
            if not date_str or not date_str[:2].isdigit():
                continue
            try:
                post_date = datetime.strptime(date_str, "%d/%m/%Y")
                if post_date >= cutoff_social:
                    row["Status"] = "APERTO"
                    row["Note aggiuntive"] = (
                        f"Promosso ad APERTO: post {col.split()[-1]} del {date_str}"
                    )
                    promoted += 1
                    break
            except:
                pass
    if promoted:
        print(f"  {city}: {promoted} ristoranti promossi INCERTO→APERTO per post social recente")
        stats[city]["aperto"]  += promoted
        stats[city]["incerto"] -= promoted
        promoted_total         += promoted
if promoted_total:
    print(f"Totale promossi: {promoted_total}")
# ══════════════════════════════════════════════════════════════════════════
# 4. SALVA EXCEL CON COLORI
# ══════════════════════════════════════════════════════════════════════════
for city in ["Milano","Roma","Torino"]:
    verif_file = os.path.join(DATA_DIR, f"Verifica_{city}_2027.xlsx")
    df_out     = pd.DataFrame(rows_by_city[city])
    df_out.to_excel(verif_file, index=False)
    wb = openpyxl.load_workbook(verif_file)
    ws = wb.active
    yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    red    = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    st_col = next((c.column for c in ws[1] if c.value == "Status"), None)
    if st_col:
        for r in ws.iter_rows(min_row=2, max_row=ws.max_row):
            sv = r[st_col-1].value
            if sv == "INCERTO":
                for c in r: c.fill = yellow
            elif sv == "CHIUSO":
                for c in r: c.fill = red; c.font = Font(color="FFFFFF")
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len+2, 40)
    wb.save(verif_file)
    print(f"  Salvato: Verifica_{city}_2027.xlsx | Social verificati: {stats[city]['social_checked']}")
# ══════════════════════════════════════════════════════════════════════════
# 5. RIEPILOGO
# ══════════════════════════════════════════════════════════════════════════
print("\n=== RIEPILOGO ===")
for city, s in stats.items():
    print(f"{city}: {s['aperto']} APERTO, {s['chiuso']} CHIUSO, {s['incerto']} INCERTO (tot {s['totale']})")
    print(f"  Mantenuti:{s['mantenuti']} Confermati:{s['verificati_oggi']} Social:{s['social_checked']}")
# ══════════════════════════════════════════════════════════════════════════
# 6. EMAIL VIA BREVO
# ══════════════════════════════════════════════════════════════════════════
print("\n=== INVIO EMAIL ===")
if not BREVO_API_KEY:
    print("BREVO_API_KEY non impostata.")
else:
    attachments = []
    for city in ["Milano","Roma","Torino"]:
        path = os.path.join(DATA_DIR, f"Verifica_{city}_2027.xlsx")
        with open(path,"rb") as f:
            b64 = base64.b64encode(f.read()).decode()
        attachments.append({"name": f"Verifica_{city}_2027.xlsx", "content": b64})
    body = (f"Verifica apertura ristoranti del {today_str}\n\n"
            + "\n".join(f"{c}: {s['aperto']} APERTO, {s['chiuso']} CHIUSO, {s['incerto']} INCERTO"
                        for c,s in stats.items())
            + f"\n\nSocial verificati: {sum(s['social_checked'] for s in stats.values())} / {len(handles_needed)} totali"
            + "\n\nIn allegato i file Excel.")
    try:
        resp = requests.post("https://api.brevo.com/v3/smtp/email",
            headers={"api-key": BREVO_API_KEY, "content-type": "application/json"},
            json={"sender": {"name":"La Pecora Nera - Sistema automatico","email":SENDER_EMAIL},
                  "to": RECIPIENTS,
                  "subject": f"Verifica apertura locali guide del {today_str}",
                  "textContent": body, "attachment": attachments},
            timeout=60)
        print(f"Email {'inviata!' if resp.status_code in (200,201) else 'ERRORE: '+str(resp.status_code)}")
    except Exception as e:
        print(f"Eccezione email: {e}")
# ══════════════════════════════════════════════════════════════════════════
# 7. DROPBOX SYNC (OUT) — ricarica i Verifica_*.xlsx aggiornati sul cloud
# ══════════════════════════════════════════════════════════════════════════
print("\n=== DROPBOX SYNC (OUT) ===")
if DROPBOX_APP_KEY and DROPBOX_APP_SECRET and DROPBOX_REFRESH_TOKEN:
    # Rigeneriamo l'access token — quello iniziale potrebbe essere scaduto
    # dopo un run lungo (la durata massima è 4 ore).
    dbx_token = dropbox_get_access_token()
    if dbx_token:
        for city in ["Milano", "Roma", "Torino"]:
            fname  = f"Verifica_{city}_2027.xlsx"
            local  = os.path.join(DATA_DIR, fname)
            remote = f"{DROPBOX_FOLDER}/{fname}"
            if dropbox_upload(dbx_token, local, remote):
                print(f"  {fname} caricato su Dropbox")
            else:
                print(f"  {fname} NON caricato su Dropbox")
    else:
        print("  Token Dropbox non ottenuto — skip upload")
else:
    print("  Dropbox non configurato — skip upload")
# ══════════════════════════════════════════════════════════════════════════
# 8. GIT PUSH
# ══════════════════════════════════════════════════════════════════════════
print("\n=== GIT PUSH ===")
if not GITHUB_TOKEN or not GITHUB_REPO:
    print("Token/repo non impostati — skip.")
else:
    try:
        subprocess.run(["git","config","user.email",SENDER_EMAIL],    cwd=SCRIPT_DIR, check=True)
        subprocess.run(["git","config","user.name","PNE Routine Bot"], cwd=SCRIPT_DIR, check=True)
        for city in ["Milano","Roma","Torino"]:
            subprocess.run(["git","add",f"Verifica_{city}_2027.xlsx"], cwd=SCRIPT_DIR, check=True)
        diff = subprocess.run(["git","diff","--cached","--quiet"], cwd=SCRIPT_DIR)
        if diff.returncode != 0:
            subprocess.run(["git","commit","-m",f"Verifica {today_str}"], cwd=SCRIPT_DIR, check=True)
            subprocess.run(["git","remote","set-url","origin",
                            f"https://{GITHUB_TOKEN}@github.com/{GITHUB_REPO}.git"],
                           cwd=SCRIPT_DIR, check=True)
            subprocess.run(["git","push"], cwd=SCRIPT_DIR, check=True)
            print("File Verifica aggiornati nel repository.")
        else:
            print("Nessuna modifica da committare.")
    except subprocess.CalledProcessError as e:
        print(f"Errore git: {e}")
print("\n=== COMPLETATO ===")
